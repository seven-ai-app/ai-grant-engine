"""Generate budget Excel spreadsheet per Innovation Authority format."""

import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..graph.state import GrantState
from ..tools.hebrew_utils import BUDGET_CATEGORIES_HE
from ..config.constants import GRANT_RATE, MAX_BUDGET_NIS, MAX_GRANT_NIS

logger = logging.getLogger(__name__)


def generate_budget_xlsx(state: GrantState, output_dir: Path) -> Path:
    """Generate budget spreadsheet in Innovation Authority format."""
    wb = Workbook()

    # Main budget sheet
    ws = wb.active
    ws.title = "תקציב מפורט"
    ws.sheet_view.rightToLeft = True
    _build_budget_sheet(ws, state)

    # Summary sheet
    ws_summary = wb.create_sheet("סיכום תקציבי")
    ws_summary.sheet_view.rightToLeft = True
    _build_summary_sheet(ws_summary, state)

    # Quarterly breakdown
    ws_quarterly = wb.create_sheet("פירוט רבעוני")
    ws_quarterly.sheet_view.rightToLeft = True
    _build_quarterly_sheet(ws_quarterly, state)

    # Save
    filename = f"תקציב_{state['startup_name'].replace(' ', '_')}.xlsx"
    output_path = output_dir / filename
    wb.save(str(output_path))
    logger.info(f"Generated XLSX: {output_path}")
    return output_path


def _build_budget_sheet(ws, state: GrantState):
    """Build the detailed budget sheet."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"תקציב מפורט - {state['startup_name']}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        ("A", "מס'", 5),
        ("B", "קטגוריה", 15),
        ("C", "תיאור", 40),
        ("D", "תעריף שעתי (ש\"ח)", 15),
        ("E", "שעות", 10),
        ("F", "סכום (ש\"ח)", 15),
        ("G", "חלק מענק (ש\"ח)", 15),
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, title, width in headers:
        cell = ws[f"{col}3"]
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col].width = width

    # Data rows
    budget_lines = state.get("budget_lines", [])
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for i, line in enumerate(budget_lines, start=1):
        row = i + 3
        ws[f"A{row}"] = i
        ws[f"B{row}"] = BUDGET_CATEGORIES_HE.get(line.get("category", ""), line.get("category", ""))
        ws[f"C{row}"] = line.get("description_he", "")
        ws[f"D{row}"] = line.get("hourly_rate") or ""
        ws[f"E{row}"] = line.get("hours") or ""
        ws[f"F{row}"] = line.get("amount", 0)
        ws[f"G{row}"] = line.get("grant_portion", 0)

        # Format numbers
        ws[f"F{row}"].number_format = "#,##0"
        ws[f"G{row}"].number_format = "#,##0"

        # Borders
        for col in "ABCDEFG":
            ws[f"{col}{row}"].border = thin_border
            ws[f"{col}{row}"].alignment = Alignment(horizontal="right")

    # Totals row
    total_row = len(budget_lines) + 4
    ws[f"C{total_row}"] = "סה\"כ"
    ws[f"C{total_row}"].font = Font(bold=True)
    ws[f"F{total_row}"] = state.get("total_budget", 0)
    ws[f"F{total_row}"].font = Font(bold=True)
    ws[f"F{total_row}"].number_format = "#,##0"
    ws[f"G{total_row}"] = state.get("grant_amount", 0)
    ws[f"G{total_row}"].font = Font(bold=True)
    ws[f"G{total_row}"].number_format = "#,##0"


def _build_summary_sheet(ws, state: GrantState):
    """Build the budget summary sheet."""
    ws["A1"] = "סיכום תקציבי"
    ws["A1"].font = Font(size=14, bold=True)

    summary_data = [
        ("תקציב כולל מבוקש", state.get("total_budget", 0)),
        ("שיעור מענק", f"{GRANT_RATE * 100:.0f}%"),
        ("מענק מבוקש", state.get("grant_amount", 0)),
        ("מימון עצמי", state.get("self_funding", 0)),
        ("", ""),
        ("תקרת תקציב מסלול", MAX_BUDGET_NIS),
        ("תקרת מענק מסלול", MAX_GRANT_NIS),
    ]

    for i, (label, value) in enumerate(summary_data, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value
        if isinstance(value, (int, float)):
            ws[f"B{i}"].number_format = "#,##0"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    # Category breakdown
    ws[f"A{len(summary_data) + 5}"] = "פירוט לפי קטגוריה"
    ws[f"A{len(summary_data) + 5}"].font = Font(size=12, bold=True)

    category_totals = {}
    for line in state.get("budget_lines", []):
        cat = line.get("category", "other")
        category_totals[cat] = category_totals.get(cat, 0) + line.get("amount", 0)

    row = len(summary_data) + 7
    for cat, total in category_totals.items():
        ws[f"A{row}"] = BUDGET_CATEGORIES_HE.get(cat, cat)
        ws[f"B{row}"] = total
        ws[f"B{row}"].number_format = "#,##0"
        row += 1


def _build_quarterly_sheet(ws, state: GrantState):
    """Build quarterly budget breakdown."""
    ws["A1"] = "פירוט רבעוני"
    ws["A1"].font = Font(size=14, bold=True)

    headers = ["פעילות", "רבעון 1", "רבעון 2", "רבעון 3", "רבעון 4", "סה\"כ"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
        ws.cell(row=3, column=i).font = Font(bold=True)

    ws.column_dimensions["A"].width = 35
    for col in "BCDEF":
        ws.column_dimensions[col].width = 12

    # Distribute budget lines across quarters (simple 4-way split for now)
    budget_lines = state.get("budget_lines", [])
    for i, line in enumerate(budget_lines, start=1):
        row = i + 3
        amount = line.get("amount", 0)
        ws.cell(row=row, column=1, value=line.get("description_he", ""))

        # Simple distribution: 30% Q1, 30% Q2, 25% Q3, 15% Q4
        q_dist = [0.30, 0.30, 0.25, 0.15]
        for q, pct in enumerate(q_dist, start=2):
            cell = ws.cell(row=row, column=q, value=round(amount * pct, -2))
            cell.number_format = "#,##0"

        ws.cell(row=row, column=6, value=amount)
        ws.cell(row=row, column=6).number_format = "#,##0"

    # Totals
    total_row = len(budget_lines) + 4
    ws.cell(row=total_row, column=1, value="סה\"כ")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    total = state.get("total_budget", 0)
    for q, pct in enumerate([0.30, 0.30, 0.25, 0.15], start=2):
        cell = ws.cell(row=total_row, column=q, value=round(total * pct, -2))
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total)
    ws.cell(row=total_row, column=6).number_format = "#,##0"
    ws.cell(row=total_row, column=6).font = Font(bold=True)
